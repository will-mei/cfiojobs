#!/usr/bin/env python
# coding=utf-8

import re 
import os
import sys
import json 
import cfiojobsutil.utils as cu
from openpyxl import Workbook
from openpyxl.chart import (
        SurfaceChart,
        ScatterChart,
        BubbleChart,
        SurfaceChart3D,
        Reference,
        Series,
)
#from openpyxl.chart.axis import SeriesAxis
from openpyxl.chart._3d import View3D
# font 
#from openpyxl.styles import Font 
#from openpyxl.drawing.text import Paragraph,ParagraphProperties,CharacterProperties,LineProperties,Font
# turn x axis line off 
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
#from openpyxl.chart.axis import ChartLines
#from openpyxl.chart.label import DataLabelList

#print(sys.argv)
# column nuber(Natural number)
#key_column_num   = 3
#data_columns     = [8,10,11,12]
#surface_columns  = map(lambda x: x , range(19,34))

# load config file 
conf_file = sys.argv[0] + '.ini'
if os.path.exists(conf_file) and os.path.isfile(conf_file):
    pass
else:
    print('config file error!')
    sys.exit(1)

def load_conf(conf_file):
    with open(conf_file,'r') as cf: 
        try:
            cf_dict = json.load(cf)
        except ValueError:
            print('warning: load json config file',cf,' failed, abort!')
            sys.exit(1)
    return cf_dict 

try:
    conf_dict = load_conf(conf_file)
    #print(conf_dict)
    key_column_num  = conf_dict['global options']['key_column_num']
    data_columns    = conf_dict['global options']['data_columns']
    surf_col_range  = conf_dict['global options']['surface_columns_range']
except:
    print('config file format error!')
    print(key_column_num,data_columns,surf_col_range)
    sys.exit(1)

key_column_index = key_column_num -1
key_column       = cu.num2letter(key_column_num)

try:
    outputfile       = './' + cu.get_file_name(sys.argv[1])[1][0:30] + '-ScatterChart' + '.xlsx'
except:
    print'Usage:', sys.argv[0],'[rbd report csv file1 file2 file3 ...]'
    sys.exit(1)

# create a Workbook for new report
wb  = Workbook()
ws  = wb.active 
ws.title = 'ScatterChart'
wb.create_sheet(title='SurfaceChart', index=0)


def get_uniq_column_info(col_name,work_sheet):
    #skip title 
    start_cell = col_name + '2'
    end_cell   = col_name + str(work_sheet.max_row)
    cell_range = str(start_cell) + ':' + str(end_cell)
#    print(cell_range)
    pattern_name_row_range = dict()
    for t in work_sheet[cell_range]:
        for c in t:
            if c.value in pattern_name_row_range.keys() :
                pattern_name_row_range[c.value].append(c.row)
            else :
                pattern_name_row_range[c.value] = list()
                pattern_name_row_range[c.value].append(c.row)
    return pattern_name_row_range 

def get_pattern_rows_map(data_work_book):
    tmp_dict = dict()
    #tmp_dict = {4k:{'Sheet1':[2,31]}}
    #            {pattern_name:{work_sheet:[line number set]}}
    #get data info of every patter in every sheet of work book
    for work_sheet in data_work_book.worksheets :
        if work_sheet.title in ['Lateral Contrast LineChart', 'ScatterChart','SurfaceChart']:
            #print(work_sheet.title)
            continue 
        #print(work_sheet.title)
        # get pattern info in sheets combined together. 
        pattern_rows_range_dict = get_uniq_column_info(key_column,work_sheet)
        for pattern_name in pattern_rows_range_dict.keys():
            if tmp_dict.has_key(pattern_name):
                # save pattern info in this sheet 
                tmp_dict[pattern_name][work_sheet.title] = pattern_rows_range_dict[pattern_name]
            else:
                tmp_dict[pattern_name] = dict()
                tmp_dict[pattern_name][work_sheet.title] = pattern_rows_range_dict[pattern_name]
    return tmp_dict 


# plot ScatterChart
def draw_pattern_scatterchart(data_work_sheet, data_rows_seq, data_columns_list, chart_position, chart_title):
    chart = ScatterChart()
    chart.title = str(chart_title)
    chart.legend.position = 't' 
    #chart.x_axis.title = 'samples from:' + data_work_sheet.title[0:21]
    #chart.x_axis.title = u'样本:' + data_work_sheet.title[0:25] + '..'
    chart.x_axis.title = data_work_sheet.title + '..'
    #chart.x_axis.txPr =
    #chart_title_font = Font(name='Arial Narrow',size=12)
    chart.y_axis.title = 'lantency(ms)'
    #chart.style = 13
    #chart.layout = 
    #chart.graphical_properties = 
    #chart.varyColors = 'blue' 
    #chart.scatterStyle = 'marker'
    #same sheet same sample width of chart, x ray values, get min and max row from pattern_col 
    xvalues = Reference(data_work_sheet, min_col=1, min_row=data_rows_seq[0], max_row=data_rows_seq[-1])
    for i in data_columns_list :
        #print(data_work_sheet.cell(row=1,column=i).value)
        line_title = data_work_sheet.cell(row=1,column=i).value
        yvalues = Reference(data_work_sheet, min_col=i, min_row=data_rows_seq[0], max_row=data_rows_seq[-1])
        series  = Series(yvalues, xvalues, title=line_title)
        chart.series.append(series)
    #print(chart.series[0])
    # Style the lines
    for i in range(len(data_columns_list)) :
        s1 = chart.series[i]
        s1.marker.symbol = "circle"
        #s1.marker.symbol = "triangle"
        #s1.marker.graphicalProperties.solidFill = "FF0000" # Marker filling
        #s1.marker.graphicalProperties.line.solidFill = "FF0000" # Marker outline
        s1.graphicalProperties.line.noFill = True 
#    print(wb.chartsheets)
    wb['ScatterChart'].add_chart(chart, chart_position)

# plot SurfaceChart for a work sheet 
def draw_pattern_surfacechart(data_work_sheet, data_rows_seq, data_columns_list, chart_position, chart_title, surf_col_range):
    # chart format 
    chart = SurfaceChart3D()
    #chart = SurfaceChart()
    chart.height = 11.5 
    chart.wireframe = True
    chart.legend.position = 'r'
    chart.title = pattern_name
    #chart.x_axis.title = 'percentile:' + data_work_sheet.title[0:25] + '...'
    chart.x_axis.title = data_work_sheet.title + '..'
    chart.y_axis.title = 'latency(ms)'

    # change view3D property 
    v3d = View3D(rotX=0.0, rotY=330L, depthPercent=120L, rAngAx=False, perspective=15L)
    #v3d.depthPercent = 20L
    chart.view3D = v3d 

    # turn majorGridlines off using shapes.GraphicalProperties and drawing.LineProperties
    #chart.y_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
    #chart.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill = '000000')
    #chart.x_axis.majorGridlines = ChartLines()
    #chart.x_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
    #chart.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill = '000000')
    #chart.dLbls = DataLabelList()
    #chart.dLbls.showVal = 0

    # add data to chart 
    pattern_name_row_range = pattern_rows_range_dict[pattern_name]
    #print(pattern_name, pattern_name_row_range[0], pattern_name_row_range[-1])

    # chart inner labels
    labels   = Reference(data_work_sheet, min_col=surf_col_range[0], max_col=surf_col_range[-1], min_row=1, max_row=1)
    # data area
    data_ref = Reference(data_work_sheet, min_col=surf_col_range[0], max_col=surf_col_range[-1], min_row=pattern_name_row_range[0], max_row=pattern_name_row_range[-1])
    #chart.add_data(data_ref, from_rows=True, titles_from_data=True) #this will lead to a bug when there are too manay rows tobe add 
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels)

    wb['SurfaceChart'].add_chart(chart, chart_position)

# plot BubbleChart 
def draw_pattern_bubblechart(data_work_sheet, data_rows_seq, data_columns_list, chart_position, chart_title):
    chart = BubbleChart()
    chart.title = str(chart_title)
    chart.style = 18
    chart.x_axis.title = data_work_sheet.title + '..'
    chart.y_axis.title = 'lantency(ms)'
    #same sheet same sample width of chart
    xvalues = Reference(data_work_sheet, min_col=1, min_row=data_rows_seq[0], max_row=data_rows_seq[-1])
    for i in data_columns_list :
        BubbleChart_line_title = data_work_sheet.cell(row=1,column=i).value
        yvalues = Reference(data_work_sheet, min_col=i, min_row=data_rows_seq[0], max_row=data_rows_seq[-1])
        size    = Reference(data_work_sheet, min_col=i, min_row=data_rows_seq[0], max_row=data_rows_seq[-1])
        series  = Series(values=yvalues, xvalues=xvalues, zvalues=size, title=BubbleChart_line_title)
        chart.series.append(series)
    wb['ScatterChart'].add_chart(chart, chart_position)

# plot ScatterChart for Lateral contrast between work sheets
def draw_lateral_pattern_scatterchart(data_work_book, data_columns_list):
    # build a new work sheet for contrast charts 
    data_work_book.create_sheet(title='Lateral Contrast LineChart', index=0)
    tmp_dict = get_pattern_rows_map(data_work_book)
    #print(tmp_dict['4kread'].keys())
    # plot chart 
    column_num = 0
    for data_column in data_columns_list:
        #column_position = cu.num2letter(column_num *8 +1)
        column_position = cu.num2letter(column_num *9 +1)
        column_num = column_num +1
        # get pattern info in sheets combined together. 
        pattern_num = 0 
        for pattern_name in cu.bp_sort(tmp_dict.keys(), screening=True):
            # 4mwrite
            #row_position = str(pattern_num *16 +1)
            #row_position = str(pattern_num *22 +1)
            row_position = str(pattern_num *26 +1)
            pattern_num = pattern_num +1
            chart_position = column_position + row_position
            #print(chart_position)
            # chart format 
            chart = ScatterChart()
            #chart.height = 10 
            chart.height = 12 
            chart.width  = 17 
            chart.title = str(pattern_name)
            chart.legend.position = 't'
            tmp_sheet = tmp_dict[pattern_name].keys()[0]
            chart.x_axis.title = wb[tmp_sheet][str(cu.num2letter(data_column)) + '1'].value 
            chart.y_axis.title = 'latency(ms)'
            # turn majorGridlines off using shapes.GraphicalProperties and drawing.LineProperties
            #chart.y_axis.majorGridlines.spPr = GraphicalProperties(noFill = 'True')
            #chart.y_axis.majorGridlines.spPr.ln = LineProperties(solidFill = '000000')
            #chart.x_axis.majorGridlines = ChartLines()
            chart.x_axis.majorGridlines.spPr = GraphicalProperties(noFill=True)
            chart.x_axis.majorGridlines.spPr.ln = LineProperties(solidFill = 'F0F0F0')
            #chart.dLbls = DataLabelList()
            #chart.dLbls.showVal = 0
            # add info from different sheet for a certain pattern , 'sheet1':[n,n+1]
            line_set_info = tmp_dict[pattern_name]
            #print(line_set_info)
            for sheetN_set_name in line_set_info.keys():
                line_title = str(sheetN_set_name)
                line_set = line_set_info[sheetN_set_name]
                #print(sheetN_set_name,line_set)
                # width (samples name)
                xvalues = Reference(data_work_book[sheetN_set_name], min_col=1, min_row=line_set[0], max_row=line_set[-1])
                # height (value point)
                yvalues = Reference(data_work_book[sheetN_set_name], min_col=data_column, min_row=line_set[0], max_row=line_set[-1])
                series  = Series(yvalues, xvalues, title=line_title)
                chart.series.append(series)
            wb['Lateral Contrast LineChart'].add_chart(chart, chart_position)

# draw patter ScatterChart for each sheet 
# chart layout: column  --  csv source 
#               row     --  pattern 

csv_num = 0 # column
for csv in sys.argv[1:]:

    # 1. check each csv and create work_sheet for them  
    if os.path.isfile(csv):
        pass
    else:
        print csv,': not a file!'
        continue 
    # sheet name less than 30 letters 
    tmp_sheet_name = cu.get_file_name(csv)[1][0:30]
    #print(cu.load_csv(csv))
    wb.create_sheet(title=tmp_sheet_name)

    # 2. load data form csv and add data(rows) into Worksheet
    for row in cu.load_csv(csv, sort_sheet=True, sort_column_index=key_column_index):
        wb[tmp_sheet_name].append(row)

    # 3. each pattern of test get its own chart, use pattern name as the chart title 
    # get pattern_rows_range_dict will calcaculate the column range for each pattern, those columns containing the data for chart to be drawing
    pattern_rows_range_dict = get_uniq_column_info(key_column, wb[tmp_sheet_name])
    #print(pattern_rows_range_dict.keys())

    # 4. calculate chart position in an individual work_sheet
    # default ScatterChart scale in unit of cell: width 8, height 16 / 25
    chart_column_position   = cu.num2letter(csv_num * 8 + 1) 
    # chart position: chart column num for current csv file 
    csv_num = csv_num + 1

    # chart position: chart row num for current pattern of data
    chart_num = 0 # row

    # 5. drawing ScatterChart
    #print(cu.bp_sort(pattern_rows_range_dict.keys(), screening=True))
    for pattern_name in cu.bp_sort(pattern_rows_range_dict.keys(), screening=True):
        # count chart number
    #    chart_num=len(wb.chartsheets) * 10 + 1

        # chart default width(8 columns) 15 height 7.5(16cells) 
        chart_row_position_scatter  = str(chart_num * 16 + 1)
        chart_row_position_surface  = str(chart_num * 25 + 1)
        chart_num = chart_num + 1

        # position str looks like this: "A1"
        chart_position_scatter  = chart_column_position + chart_row_position_scatter 
        chart_position_surface  = chart_column_position + chart_row_position_surface 
    #    print(position)

        # plot chart
        draw_pattern_scatterchart(wb[tmp_sheet_name], pattern_rows_range_dict[pattern_name], data_columns, chart_position_scatter, pattern_name)
        draw_pattern_surfacechart(wb[tmp_sheet_name], pattern_rows_range_dict[pattern_name], data_columns, chart_position_surface, pattern_name, surf_col_range)


if len(sys.argv[1:]) > 1 :
    draw_lateral_pattern_scatterchart(wb, data_columns) 

wb.save(outputfile)
