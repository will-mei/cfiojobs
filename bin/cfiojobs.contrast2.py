#!/usr/bin/env python
# -*- coding:utf-8 -*-

from __future__ import print_function
import cfiojobsutil.utils as cu
import sys
# openpyxl packages
from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
try:
    parallel_csv    = sys.argv[1]
    single_csv      = sys.argv[2]
#    output_file     = './test.xlsx'
except IndexError:
    print("this script contrast fio parallel mode test report with single mode report, default output file name: './<your csv file name>-report.xlsx'.\nand it requires: \n\t1. parallel mode report csv file \n\t2. single mode report csv file \n3. disk type \n4. excel name :xxx.xlsx \nplease pass in these names first.")
    exit()

try:
    n= '-' + sys.argv[3] + '-report'
except:
    n='-report'
    pass 

# if output file name spcified
try:
    output_file = sys.argv[4]
except:
    output_file = cu.get_file_name(parallel_csv)[1][0:30].split('_')[0] + n + '.xlsx'
# if output file an existing excel file 
try:
    # load and create a new 
    wb = load_workbook(output_file)
    wb.create_sheet(title='tmp_sheet', index=0)
    ws = wb['tmp_sheet']
except:
    wb = Workbook()
    ws = wb.active 

reference_index = 0.9
key_index_list  = [0,1,2]

#err_stat   = u'…'.encode('GB2312')
ok_stat    = u'○'
bad_stat   = u'●'
err_stat   = u'…'
stat_pool  = [ok_stat,bad_stat,err_stat]

# parallel first
def tweens_report(dic_p, dic_s, ref_index, stat_list):
    result_array = list()
    # get parallel mode info
    #for i in sorted(dic_p.keys(), reverse=True):
    #for i in sorted(dic_p.keys())[-1:] and sorted(dic_p.keys())[:-1]: 
    #for i in sorted(dic_p.keys())[-1:] + sorted(dic_p.keys())[:-1]:
    k = sorted(dic_p.keys())
    for i in k[-1:] + k[:-1]:
        row = list()
        # get single mode info
        if dic_s.has_key(i):
            row += dic_s[i][0:7]
            row += dic_p[i][3:6]
            try :
                v_ref = dic_p[i][3] / dic_s[i][4]
                if v_ref > ref_index :
                    #continue 
                    stat = stat_list[0]
                else:
                    stat = stat_list[1]
                row.append(str(round(v_ref * 100 ,2)) + '%')
            except:
                stat = stat_list[-1]
                row.append(stat)
            row.append(stat)
            result_array.append(row)
    return result_array 


# load data form csv
single_report   = cu.load_csv(single_csv)
d= {13:'iodepth',
    14:'numjobs',
    16:'size',
    17:'runtime',
    18:'ioengine'
   }
job_arg_info = ''.join(map(lambda x: str(d[x]) + '=' + str(int(single_report[1][x])) + '  ' if isinstance(single_report[1][x], float) else str(d[x]) + '=' + str(single_report[1][x]) + '  ', sorted(d.keys()) ))


# check disk number (include title)
disk_count = len(set(map(lambda x: x[1], single_report)))

if disk_count > 2:
    parallel_report = cu.load_csv(parallel_csv)
    #single_dict     = cu.array2dic(single_report, key_index_list)
    #parallel_dict   = cu.array2dic(parallel_report, key_index_list)
    single_dict     = cu.report_to_sortable_dic(single_report)
    parallel_dict   = cu.report_to_sortable_dic(parallel_report)
    # contrast paralle to single
    result_sheet = tweens_report(parallel_dict, single_dict, reference_index, stat_pool)
    #print(result_sheet)
else:
    result_sheet = map(lambda x : x[0:7], single_report)

#print(map(lambda x : cu.bp2num(x[2]), result_sheet[0:10]))

####################################################################
#wb=Workbook()
#ws=wb.active

#get host number - title 
h_list   = map(lambda x: x[0], result_sheet)
u_list   = {}.fromkeys(h_list).keys()
h_count  = len(u_list) -1 

# give a new name of worksheet
tmp_sheet_name = cu.get_file_name(single_csv)[1][0:30].split('_')[0] + '(' + str(h_count) + u'台' +  ')'
# check name confliction in worksheet 
if tmp_sheet_name in wb.sheetnames:
    same_name_num=map(lambda x : tmp_sheet_name in x.title , wb).count(True)
    tmp_sheet_name = tmp_sheet_name + str(same_name_num)
    ws.title = tmp_sheet_name 
else:
    ws.title = tmp_sheet_name 

# add sheet to Workbook
for row in result_sheet:
        ws.append(row)

# set title format 
column_max = cu.num2letter(ws.max_column)
title_cell = 'A1:' + column_max + '1'
#print(title_cell)

####################################################################
title_dict = {
'A1':'主机名',
'B1':'测试设备',
'C1':'块大小/模式',
'D1':'该盘单盘独立测试\n带宽值(MiB/s)',
'E1':'单盘独立测试的\n平均带宽值(MiB/s)',
'F1':'该盘单盘独立测试带宽值/\n单盘独立测试的平均带宽值(%)',
'G1':'单测筛选状态',
'H1':'该盘并行测试时\n带宽值(MiB/s)',
'I1':'并行测试的\n平均带宽值(MiB/s)',
'J1':'该盘并行测试时的带宽值/\n并行测试的平均带宽值(%)',
'K1':'该盘并行测试时带宽值/\n单盘独立测试的平均带宽值(%)',
'L1':'最终筛选状态'
}
for k in title_dict:
    if k[0] <= column_max:
        ws[k].value = title_dict[k]

####################################################################
# width = Chinese characters number * 2
# width = ascii characters number * 1.1
line_width_dict = {
    'A':14.38,
#    'B':8.88,
    'B':13.25,
#    'c':11.5,
    'C':14.38,
    'D':16.63,
    'E':18.13,
    'F':28.25,
    'G':12.38,
    'H':14.5,
    'I':18.13,
    'J':24.13,
    'K':28.25,
    'L':13.5
}
for k in line_width_dict:
    if k[0] <= column_max:
        ws.column_dimensions[k].width = line_width_dict[k]


####################################################################
# set title fill 
title_fill = PatternFill("solid", fgColor="FFA500")

# set title border 
bd       = Side(style='thin', color="000000")
bd_round = Border(left=bd, top=bd, right=bd, bottom=bd)

#highlight = NamedStyle(name="highlight")
#highlight.font = Font(bold=True, size=12)
#highlight.font = Font(bold=True)
#highlight.border = bd_round

#for cell in ws['A1:L1'][0]:
for cell in ws[title_cell][0]:
#    print(cell) 
    cell.fill = title_fill
    cell.border = bd_round
    cell.alignment = Alignment(wrap_text=True)
    #cell.alignment = Alignment(wrap_text=False, shrink_to_fit=False, indent=0)


####################################################################
bold_font = Font(bold=True)

font_dict = {
    'L1':bold_font
}
for k in font_dict:
    if k[0] <= column_max:
        ws[k].font = font_dict[k]

####################################################################
# must be unicode 
comment_dict = {
    'C1':job_arg_info,
    'E1':u'同一时间同一规格主机上\n测试一块磁盘在某一模式的读写测试，\n所有同批次进行单盘测试主机的\n所有同类型磁盘的带宽平均值',
    'G1':u'该盘单盘独立测试带宽值/\n单盘独立测试的平均带宽值;\n低于90%标为●否则记为○',
    'I1':u'同一时间同一规格主机上\n所有数据磁盘进行同一模式的读写测试,\n所有同批次进行并行测试主机的\n所有同类型磁盘的带宽平均值',
    'L1':u'该盘并行测试带宽值/\n单盘独立测试的平均带宽值;\n低于90%标为●否则记为○'
}
for k in comment_dict:
    if k[0] <= column_max:
        t         = comment_dict[k]
        cmt       = Comment(t, 'note')
        cmt.width = 300
        ws[k].comment = cmt 

####################################################################
# freeze the first line 
ws.freeze_panes = 'A2' 

####################################################################

# add avg sheet 
# load csv 
s_avg_sheet_file = sys.argv[2].replace('all_host','avg')
p_avg_sheet_file = sys.argv[1].replace('all_host','avg')
#print(s_avg_sheet_file)
s_avg_sheet = sorted(cu.load_csv(s_avg_sheet_file), key=lambda x : cu.bp2num(x[0]))
p_avg_sheet = sorted(cu.load_csv(p_avg_sheet_file), key=lambda x : cu.bp2num(x[0]))
#print(s_avg_sheet)
# create worksheet
avg_sheet_name = tmp_sheet_name+'avg'
wb.create_sheet(title=avg_sheet_name, index=1)
wa = wb[avg_sheet_name]
# add to worksheet
wa.append(['single'])
for row in s_avg_sheet[-1:] + s_avg_sheet[:-1]:
    wa.append(row)
wa.append([''])
wa.append(['parallel'])
for row in p_avg_sheet[-1:] + p_avg_sheet[:-1]:
    wa.append(row)
# col width 
for k in list(map(chr, range(ord('A'), ord(get_column_letter(wa.max_column)) + 1))):
    if k[0] <= cu.num2letter(wa.max_column):
        wa.column_dimensions[k].width = 20
        wa[k+'2'].fill =title_fill 

wb.save(output_file)
