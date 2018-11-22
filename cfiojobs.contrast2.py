#!/usr/bin/env python
# -*- coding:utf-8 -*-

from __future__ import print_function
import cfiojobsutil.utils as cu
import sys
# openpyxl packages
from openpyxl import Workbook 
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill,Alignment,Border,Side
try:
    parallel_csv    = sys.argv[1]
    single_csv      = sys.argv[2]
#    output_file     = './test.xlsx'
except IndexError:
    print("this script contrast fio sing mode test report with parallel mode report, new report name: ./'test-contrast.csv'.\nand it requires: \n\t1. sing mode report csv filee \n\t2. parallel mode report csv file \n please pass in these names first.")
    exit()
try:
    n= '-' + sys.argv[3] + '-report'
except:
    n='-report'
    pass 
output_file = cu.get_file_name(single_csv)[1][0:30].split('_')[0] + n + '.xlsx'


reference_index = 0.9
key_index_list  = [0,1,2]

#ok_stat    = u'○'.encode('GB2312')
#bad_stat   = u'●'.encode('GB2312')
#err_stat   = u'…'.encode('GB2312')
ok_stat    = u'○'
bad_stat   = u'●'
err_stat   = u'…'
stat_pool  = [ok_stat,bad_stat,err_stat]


def tweens_report(dic_s, dic_p, ref_index, stat_list):
    result_array = list()
    # get parallel mode info
    for i in dic_p:
        row = list()
        # get single mode info
        if dic_s.has_key(i):
            row += dic_s[i][0:7]
            row += dic_p[i][3:6]
            try :
                v_ref = dic_p[i][3] / dic_s[i][4]
                if v_ref > ref_index :
                    #continue 
                    stat = stat_list[0]
                else:
                    stat = stat_list[1]
                row.append(str(round(v_ref * 100 ,2)) + '%')
            except:
                stat = stat_list[-1]
                row.append(stat)
            row.append(stat)
            result_array.append(row)
    return result_array 


# load data form csv
parallel_report = cu.load_csv(parallel_csv)
parallel_dict   = cu.array2dic(parallel_report, key_index_list)

single_report   = cu.load_csv(single_csv)
single_dict     = cu.array2dic(single_report, key_index_list)

# contrast paralle to single
result_sheet = tweens_report(single_dict, parallel_dict, reference_index, stat_pool)
#print(result_sheet)
result_sheet.sort(key=lambda x : x[0], reverse=True )



####################################################################
wb=Workbook()
ws=wb.active

#get host number
h_list   = map(lambda x: x[0], result_sheet)
u_list   = {}.fromkeys(h_list).keys()
h_count  = len(u_list)

#b_list   = map(lambda x: x[1], result_sheet)
#ub_list  = {}.fromkeys(b_list).keys()
#b_count  = len(ub_list)

tmp_sheet_name = cu.get_file_name(single_csv)[1][0:30].split('_')[0] + '(' + str(h_count) + u'台' +  ')'
ws.title = tmp_sheet_name 

for row in result_sheet:
        ws.append(row)

# set title format 
title_cell = 'A1:' + cu.decimal2alphabet(ws.max_column) + '1'
#print(title_cell)

####################################################################
title_dict = {
'A1':'主机名',
'B1':'测试设备',
'C1':'块大小/模式',
'D1':'该盘单盘独立测试\n带宽值(MiB/s)',
'E1':'单盘独立测试的\n平均带宽值(MiB/s)',
'F1':'该盘单盘独立测试带宽值/\n单盘独立测试的平均带宽值(%)',
'G1':'单测筛选状态',
'H1':'该盘并行测试时\n带宽值(MiB/s)',
'I1':'并行测试的\n平均带宽值(MiB/s)',
'J1':'该盘并行测试时的带宽值/\n并行测试的平均带宽值(%)',
'K1':'该盘并行测试时带宽值/\n单盘独立测试的平均带宽值(%)',
'L1':'最终筛选状态'
}
for k in title_dict:
    ws[k].value = title_dict[k]

####################################################################
# width = Chinese characters number * 2
# width = ascii characters number * 1.1
line_width_dict = {
    'c':11.5,
    'D':16.63,
    'E':18.13,
    'F':28.25,
    'G':12.38,
    'H':14.5,
    'I':18.13,
    'J':24.13,
    'K':28.25,
    'L':13.5
}
for k in line_width_dict:
    ws.column_dimensions[k].width = line_width_dict[k]


####################################################################
# set title fill 
title_fill = PatternFill("solid", fgColor="FFA500")

# set title border 
bd       = Side(style='thin', color="000000")
bd_round = Border(left=bd, top=bd, right=bd, bottom=bd)

#highlight = NamedStyle(name="highlight")
#highlight.font = Font(bold=True, size=12)
#highlight.font = Font(bold=True)
#highlight.border = bd_round

#for cell in ws[title_cell][0]:
for cell in ws['A1:L1'][0]:
#    print(cell) 
    cell.fill = title_fill
    cell.border = bd_round
    cell.alignment = Alignment(wrap_text=True)
    #cell.alignment = Alignment(wrap_text=False, shrink_to_fit=False, indent=0)


####################################################################
bold_font = Font(bold=True)

font_dict = {
    'L1':bold_font
}
for k in font_dict:
    ws[k].font = font_dict[k]

####################################################################
# must be unicode 
comment_dict = {
    'E1':u'同一时间同一规格主机上\n测试一块磁盘在某一模式的读写测试，\n所有同批次进行单盘测试主机的\n所有同类型磁盘的带宽平均值',
    'G1':u'该盘单盘独立测试带宽值/\n单盘独立测试的平均带宽值;\n低于90%标为●否则记为○',
    'I1':u'同一时间同一规格主机上\n所有数据磁盘进行同一模式的读写测试,\n所有同批次进行并行测试主机的\n所有同类型磁盘的带宽平均值',
    'L1':u'该盘并行测试带宽值/\n单盘独立测试的平均带宽值;\n低于90%标为●否则记为○'
}
for k in comment_dict:
    t         = comment_dict[k]
    cmt       = Comment(t, 'note')
    cmt.width = 300
    ws[k].comment = cmt 

wb.save(output_file)
